import openpyxl

# only use videogames sheet once testing is complete - too intense for mac
path = 'openpyxl_data.xlsx'  # 1. specify filepath(s)
workbook_object = openpyxl.load_workbook(path)  # 2. create workbook object using .load_workbook(path)
active_sheet_object = workbook_object.active  # 3. get active sheet from workbook object

# this can also be done as active = workbook_object['sheet_name']

# selected_cell_object = active_sheet_object.cell(row=1, column=1)  # 4. cell object has row, column, coordinate
#
# cell_value = f"Values extracted from the select cell object are: {selected_cell_object.value}"
# rows_value = f"Number of rows in the sheet at {path} : {active_sheet_object.max_row}"
# columns_value = f"Number of columns in the sheet at {path} : {active_sheet_object.max_column}"
#
# for row in active_sheet_object.iter_rows():
#     for cell in row:
#         print(cell.value)

# print(f'Total number of rows: {active_sheet_object.max_row}\nTotal number of cols: {active_sheet_object.max_column}')

# ws.max_column = total number of columns in the worksheet (the number of columns in the first row).
# go through each column number in the range from 1 to the maximum column number (inclusive). This means it will visit
# each column in the first row, starting from the first column (column 1) and ending at the last column.
# For each column number in the range find the value of the cell located in the first row and that particular column.
# ws.cell(row=1, column=i).value to get the value of the cell at row 1 and column i.
# The extracted values are then collected into the list called values.

values = [active_sheet_object.cell(row = i, column = 1).value for i in range(1, active_sheet_object.max_column + 1)]
print(values)

# WRITING TO CELLS
active_sheet_object['A1'] = 'Item_code'
workbook_object.save('openpyxl_data.xlsx')

